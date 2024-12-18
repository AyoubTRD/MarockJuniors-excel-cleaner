import openpyxl as xl

INFO_COLUMN_START = 2
INFO_COLUMN_END = 7

ACTIVITY_COLUMN_START = 8
ACTIVITY_COLUMN_END = 27

CONSECUTIVE_EMPTY_ROWS_TO_STOP = 5

DATA_START_ROW = 4

class ExcelManipulator:
  def __init__(self, file_path):
    self.file_path = file_path
    self._load_file()  
    
  def _load_file(self):
    self.workbook = xl.load_workbook(self.file_path)
    self.worksheet = self.workbook.get_sheet_by_name(self.get_sheet_names()[0])

  def reload_file(self):
    self._load_file()

  def get_sheet_names(self):
    return self.workbook.sheetnames
  
  def set_worksheet(self, sheet_name):
    self.worksheet = self.workbook.get_sheet_by_name(sheet_name)

  def extract_data(self):
    data = []

    row_index = DATA_START_ROW
    consecutive_empty_rows = 0

    while consecutive_empty_rows < CONSECUTIVE_EMPTY_ROWS_TO_STOP:
      row = self._get_row_by_index(row_index)

      if self._row_is_empty(row):
        consecutive_empty_rows += 1
        row_index += 1
        continue
      
      consecutive_empty_rows = 0

      data.append(row)

      row_index += 1

    return data
  
  def clear_data(self):
    existing_data = self.extract_data()

    for i in range(DATA_START_ROW, DATA_START_ROW + len(existing_data)):
      for column_index in range(INFO_COLUMN_START, INFO_COLUMN_END + 1):
        self.worksheet.cell(row=i, column=column_index).value = ''
      for column_index in range(ACTIVITY_COLUMN_START, ACTIVITY_COLUMN_END + 1):
        self.worksheet.cell(row=i, column=column_index).value = ''
    
  
  def write_new_data(self, data):
    self.clear_data()

    row_index = DATA_START_ROW
    for row in data:
      i = 0
      
      for column_index in range(INFO_COLUMN_START, INFO_COLUMN_END + 1):
        self.worksheet.cell(row=row_index, column=column_index).value = row[i]
        i += 1
      
      for column_index in range(ACTIVITY_COLUMN_START, ACTIVITY_COLUMN_END + 1):
        self.worksheet.cell(row=row_index, column=column_index).value = row[i]
        i += 1
      
      row_index += 1
    
  def save_file(self, file_path):
    xl.writer.excel.save_workbook(self.workbook, file_path)

  def _get_row_by_index(self, row_index):
    row = []

    for column_index in range(INFO_COLUMN_START, INFO_COLUMN_END + 1):
      row.append(self.worksheet.cell(row=row_index, column=column_index).value)
    
    for column_index in range(ACTIVITY_COLUMN_START, ACTIVITY_COLUMN_END + 1):
      row.append(self.worksheet.cell(row=row_index, column=column_index).value)
    
    return row

  def _row_is_empty(self, row):
    for value in row:
      if value: return False
    
    return True

if __name__ == '__main__':
  excel_file_path = '/Users/ayoub/Documents/goaluin/data-duplication-solution/tests/Example 1.xlsx'
  manipulator = ExcelManipulator(excel_file_path)

  data = manipulator.extract_data()

  print('-'*30 + ' Extracted data ' + '-'*30)
  for row in data: print(row)